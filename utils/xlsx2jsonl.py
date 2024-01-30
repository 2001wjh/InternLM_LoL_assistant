# -*- coding: gbk -*-
import openpyxl
import json


def process_excel_to_jsonl(input_file_path, output_file_path):
    # Load the Excel file
    workbook = openpyxl.load_workbook(input_file_path)
    sheet = workbook['Sheet1']

    # Process the data
    data = []

    for row in sheet.iter_rows(2, sheet.max_row + 1, values_only=True):
        input_value = str(row[0])

        Strength = '{}��Ӣ��ǿ����{}����'.format(str(row[0]), str(row[1]))

        Position = '���Ӣ��ͨ���������{}λ�ã�'.format(str(row[2]))

        Win_Rate = 'ʤ����{}��'.format(str(row[3]))

        Pick_Rate = 'ѡȡ����{}��'.format(str(row[4]))

        Ban_rate = 'banѡ����{}��'.format(str(row[5]))


        if row[6]:
            weak_against_list = row[6].replace("[", "").replace("]", "").replace("'", "").split('��')
            Weak_Against = '���Ӣ�۲��ó��Կ���' + weak_against_list[-1] + '��'

        else:
            Weak_Against = '���Ӣ�۵����������Ϣ������'

        Hero_Url = '��������˽�{}����ϸ�츳�ͳ�װ��Ϣ����鿴������ַ{}��'.format(str(row[0]), str(row[7]))

        output_value = "".join([Strength, Position, Win_Rate, Pick_Rate, Ban_rate, Weak_Against, Hero_Url])

        conversation = {
            "system": "You are a professional LoL player, experienced professor, and world champion. "
                      "You always provide accurate, comprehensive and detailed answers to players' questions",
            "input": '��ѡ���Ӣ����{}����������Ӣ���ڵ�ǰ�汾�ľ�����Ϣ'.format(input_value),
            "output": output_value
        }
        data.append({"conversation": [conversation]})

    # Convert to JSONL format
    jsonl_data = '\n'.join(json.dumps(entry, ensure_ascii=False) for entry in data)

    # Save the output
    with open(output_file_path, 'w', encoding='gbk') as file:
        file.write(jsonl_data)

# Example usage
input_file_path = 'champions_data.xlsx'  # Path to your Excel file
output_file_path = 'processed_champions_data_gbk.jsonl'  # Desired output file path

process_excel_to_jsonl(input_file_path, output_file_path)


def convert_gbk_to_utf8(gbk_file_path, utf8_file_path):
    with open(gbk_file_path, 'r', encoding='gbk') as gbk_file, \
         open(utf8_file_path, 'w', encoding='utf-8') as utf8_file:
        for line in gbk_file:
            # ֱ��д��UTF-8�������
            utf8_file.write(line)

# ���ú�������ת��
gbk_file_path = 'processed_champions_data.jsonl'  # �滻Ϊ����GBK�ļ�·��
utf8_file_path = 'processed_champions_data_utf8.jsonl'  # �滻Ϊ��ϣ��������UTF-8�ļ�·��

convert_gbk_to_utf8(gbk_file_path, utf8_file_path)


